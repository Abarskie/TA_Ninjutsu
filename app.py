from flask import Flask, render_template, request, redirect, url_for, send_file
import csv
import os
import logging
import webbrowser
import threading
import time
from datetime import datetime
from openpyxl import Workbook
from flask import send_file
import io
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from google.oauth2.service_account import Credentials
import json
from dotenv import load_dotenv


app = Flask(__name__)
load_dotenv()
# Google Sheets setup
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
ASIN_SHEET_ID = "1YfkLmtVpCxnW351gwVjVnodXyOHcOUlHtCKlhFC_Cpg"
GATED_BRANDS_SHEET_ID = "1NJ0lBWNKgmtk1XX563e-E7z7nZpK17mr4cWEiu0ACco"

def get_google_credentials():
    """Load Google credentials from environment variables"""
    creds_dict = {
        "type": "service_account",
        "project_id": os.getenv("GOOGLE_PROJECT_ID"),
        "private_key_id": os.getenv("GOOGLE_PRIVATE_KEY_ID"),
        "private_key": os.getenv("GOOGLE_PRIVATE_KEY").replace('\\n', '\n'),
        "client_email": os.getenv("GOOGLE_CLIENT_EMAIL"),
        "client_id": os.getenv("GOOGLE_CLIENT_ID"),
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
        "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
        "client_x509_cert_url": os.getenv("GOOGLE_CLIENT_X509_CERT_URL")
    }
    return Credentials.from_service_account_info(creds_dict, scopes=SCOPES)

def save_asins_to_google_sheets(asins, brands, skip_brands=False):
    try:
        creds = get_google_credentials()
        client = gspread.authorize(creds)

        gated_brands = []
        if skip_brands:
            gated_sheet = client.open_by_key(GATED_BRANDS_SHEET_ID).sheet1
            gated_brands = [row[0].lower() for row in gated_sheet.get_all_values()[1:]]

        main_sheet = client.open_by_key(ASIN_SHEET_ID).sheet1
        existing_asins = [row[0] for row in main_sheet.get_all_values()[1:]]

        new_asins = []
        duplicate_count = 0
        gated_count = 0

        for asin, brand in zip(asins, brands):
            if asin in existing_asins:
                duplicate_count += 1
                continue
            if skip_brands and brand.lower() in gated_brands:
                gated_count += 1
                continue
            new_asins.append(asin)

        if new_asins:
            main_sheet.append_rows([[asin] for asin in new_asins])
            return (
                f"Added {len(new_asins)} new ASINs. "
                f"Skipped {duplicate_count} duplicates and {gated_count} gated brands."
            )
        return f"No ASINs added. Skipped {duplicate_count} duplicates and {gated_count} gated brands."

    except Exception as e:
        return f"Error: {str(e)}"

# Setup logging
LOG_FILE = "logs.txt"
logging.basicConfig(filename=LOG_FILE, level=logging.ERROR, format="%(asctime)s - %(levelname)s - %(message)s")

def log_error(error_message):
    with open(LOG_FILE, "a") as log_file:
        log_file.write(f"{datetime.now()} - ERROR - {error_message}\n")

def process_sellers(file_path):
    seller_ids = []
    
    try:
        with open(file_path, mode='r', encoding='utf-8') as file:
            content = file.read().lstrip('\ufeff')
            reader = csv.DictReader(content.splitlines())
            
            headers = reader.fieldnames or []
            required_columns = ["Seller: Review Count (Lifetime)", "Seller: ID"]
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                error_msg = f"Missing required columns: {', '.join(missing_columns)}"
                log_error(error_msg)
                os.remove(file_path)
                return error_msg
                           
            for row in reader:                                                                            
                try:
                    review_count = int(row["Seller: Review Count (Lifetime)"].strip())
                    if review_count < 500:                                                               
                        seller_ids.append(row["Seller: ID"].strip())                 
                except ValueError:
                    continue
        
        batch_size = 50
        batches = [seller_ids[i:i + batch_size] for i in range(0, len(seller_ids), batch_size)]
        return batches
    
    except Exception as e:
        log_error(f"Error processing {os.path.basename(file_path)}: {str(e)}")
        return f"Processing error: {str(e)}"
    
    finally:
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                log_error(f"Failed to delete {file_path}: {str(e)}")

def process_multiple_sellers(file_paths):
    all_seller_ids = []
    errors = []
    
    for file_path in file_paths:
        result = process_sellers(file_path)
        
        if isinstance(result, str):
            errors.append(result)
        else:
            for batch in result:
                all_seller_ids.extend(batch)
    
    batch_size = 50
    combined_batches = [all_seller_ids[i:i + batch_size] 
                       for i in range(0, len(all_seller_ids), batch_size)]
    
    return combined_batches, errors

@app.route("/", methods=["GET", "POST"])
def index():
    batches = None
    error = None

    if request.method == "POST":
        try:
            if "file" not in request.files:
                error = "No file uploaded."
                log_error(error)
                return render_template("index.html", batches=None, error=error)

            files = request.files.getlist('file')
            if not files or all(f.filename == '' for f in files):
                error = "No selected files."
                log_error(error)
                return render_template("index.html", batches=None, error=error)

            file_paths = []
            for file in files:
                if file.filename:
                    file_path = os.path.join("uploads", file.filename)
                    os.makedirs("uploads", exist_ok=True)
                    file.save(file_path)
                    file_paths.append(file_path)

            batches, processing_errors = process_multiple_sellers(file_paths)
            
            if processing_errors:
                error = " | ".join(processing_errors)
                log_error(error)

        except Exception as e:
            error = f"Unexpected error: {e}"
            log_error(error)
            batches = None

    return render_template("index.html", batches=batches, error=error)

@app.route("/clear", methods=["POST"])
def clear_batches():
    return redirect(url_for("index"))

@app.route("/import_asin", methods=["POST"])
def import_asin():
    skip_brands = request.form.get("skip_brands") == "on"
    asins = []
    brands = []
    
    for file in request.files.getlist("file"):
        content = file.read().decode("utf-8").lstrip('\ufeff')
        reader = csv.DictReader(content.splitlines())
        for row in reader:
            asins.append(row["ASIN"].strip())
            brands.append(row.get("Brand", "").strip().lower())

    skip_report = {
        "duplicate_asins": 0,
        "gated_brands": {},
        "total_skipped": 0,
        "new_asins": []
    }

    try:
        creds = get_google_credentials()
        client = gspread.authorize(creds)

        gated_brands = []
        if skip_brands:
            gated_sheet = client.open_by_key(GATED_BRANDS_SHEET_ID).sheet1
            gated_brands = [row[0].lower().strip() for row in gated_sheet.get_all_values()[1:]]

        main_sheet = client.open_by_key(ASIN_SHEET_ID).sheet1
        existing_asins = [row[0].strip() for row in main_sheet.get_all_values()[1:]]

        for asin, brand in zip(asins, brands):
            if asin in existing_asins:
                skip_report["duplicate_asins"] += 1
                continue
            
            if skip_brands and brand in gated_brands:
                skip_report["gated_brands"][brand] = skip_report["gated_brands"].get(brand, 0) + 1
                continue
                
            skip_report["new_asins"].append(asin)

        if skip_report["new_asins"]:
            main_sheet.append_rows([[asin] for asin in skip_report["new_asins"]])

        skip_report["total_skipped"] = skip_report["duplicate_asins"] + sum(skip_report["gated_brands"].values())
        report_msg = generate_skip_report(skip_report)

        return render_template("index.html",
                           asin_message=report_msg,
                           asin_data=",".join(skip_report["new_asins"]),
                           show_asin_results=True)

    except Exception as e:
        return render_template("index.html", error=f"Error: {str(e)}")

def generate_skip_report(skip_report):
    msg = f"Success! {len(skip_report['new_asins'])} new ASINs added.\n"
    
    if skip_report["duplicate_asins"] > 0:
        msg += f"\n⏩ Skipped {skip_report['duplicate_asins']} duplicate ASINs"
    
    if skip_report["gated_brands"]:
        msg += "\n\n🚫 Gated Brands Skipped:"
        for brand, count in skip_report["gated_brands"].items():
            msg += f"\n• {brand.title()}: {count} ASINs"
    
    return msg

@app.route('/download_asin_excel', methods=['POST'])
def download_asin_excel():
    try:
        asin_data = request.form.get('asin_data', '')
        asin_list = [asin.strip().strip("'\"[]") for asin in asin_data.split(',') if asin.strip()]
        
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'ASIN'
        
        for idx, asin in enumerate(asin_list, start=2):
            cell = ws.cell(row=idx, column=1, value=asin)
            cell.number_format = '@'
        
        ws.column_dimensions['A'].width = 15
        
        wb.save(output)
        output.seek(0)

        filename = f"ASIN_Export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        log_error(f"Error generating Excel file: {e}")
        return render_template("index.html", 
                           batches=None,
                           error=f"Error generating Excel file: {e}")
    
def open_browser():
    time.sleep(1)
    webbrowser.open_new_tab('http://127.0.0.1:5000')

if __name__ == "__main__":
    threading.Thread(target=open_browser).start()
    app.run(port=5000)